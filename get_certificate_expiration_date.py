from datetime import datetime, timedelta
import openpyxl
from easygui import msgbox

# Создаем объект wb, прочитав с помощью функции load_workbook() лист Excel
wb = openpyxl.load_workbook("certificate_tracking_app.xlsx")
sheet = wb["Лист1"]
sheet_certdays = wb["certdays"]
sheet_certdays_b1 = sheet_certdays["B1"].value

def get_certificate_expiration_date():
    """ Функция отображает сведения по датам окончания аттестатов"""   
    
    # Создаем переменную delta, равную количествам дней из листа certdays ячейки B1, для сравнения с ней 
    delta = timedelta(days= sheet_certdays_b1)

    # Создаем пустые списки, в которых будут аккумулироваться сообщения об аттестатах
    message_1 = []
    message_2 = []
    message_3 = [] 

    i = 0 # Общий счетчик ячеек
    j = 0 # Счетчик ячеек, когда до даты окончания аттестата больше sheet_certdays_b1 дней
    n = 0 # Счетчик объектов типа "None" (Пустые ячейки)

    # Запускаем цикл по ячейкам в столбце "D" Excel
    for cell_d in sheet["D"]:
        i +=1
        if (cell_d.value == "Дата окончания аттестата"):
            continue
        if cell_d.value == None:
            n +=1
            continue
        if (cell_d.value - datetime.now()) > delta:
            j +=1
        # Условие 1, когда у объекта истекает срок действия аттестата в течении sheet_certdays_b1 дней.
        if ((cell_d.value - datetime.now()) <= delta) and (cell_d.value > datetime.now()):
            message_1.append(f"\nВ течение {delta.days} дней у объекта {(sheet.cell(row = i, column = 2)).value} истекает срок действия аттестата {cell_d.value.strftime('%d.%m.%Y')}")
        # Условие 2, когда у объекта срок действия аттестата сегодня
        if (cell_d.value.day == datetime.now().day) and (cell_d.value.month == datetime.now().month) and (cell_d.value.year == datetime.now().year):
            message_2.append(f"\nСегодня {cell_d.value.strftime('%d.%m.%Y')} у объекта {(sheet.cell(row = i, column = 2)).value} истекает срок действия аттестата")
        # Условие 3, когда у объекта уже истек срок действия аттестата
        if (datetime.now() > cell_d.value) and not ((cell_d.value.day == datetime.now().day) and (cell_d.value.month == datetime.now().month) and (cell_d.value.year == datetime.now().year)):
            message_3.append(f"\nУ объекта {(sheet.cell(row = i, column = 2)).value} истек срок действия аттестата {cell_d.value.strftime('%d.%m.%Y')}")
    
    # Случай, когда Условия 1, 2, 3 неверны 
    if (j+n) == len(sheet["D"])-1:
        message_1.append(f"В ближайшие {delta.days} дней ни один из аттестатов не оканчивается")

    # С помощью метода join() соединяем элементы списка в строку
    changed_message_1 = " ".join(message_1) 
    changed_message_2 = " ".join(message_2) 
    changed_message_3 = " ".join(message_3) 

    # Формируем объединенное сообщение
    sum_changed_message = f"\n{changed_message_1}\n{changed_message_2}\n{changed_message_3}"

    # Возвращаем функцию msgbox модуля easygui
    return msgbox(sum_changed_message,"ЦИТСиЗИ МВД по Республике Калмыкия")
