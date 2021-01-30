from datetime import datetime, timedelta
import openpyxl
from easygui import msgbox


# Создаем объект wb, прочитав с помощью функции load_workbook() лист Excel
wb = openpyxl.load_workbook("certificate_tracking_app.xlsx")
sheet = wb["Лист1"]
sheet_certdays = wb["certdays"]

sheet_certdays_b2 = sheet_certdays["B2"].value


def get_annual_control():
    """ Функция выводит сведения по ежегодному контролю"""

    # Создаем переменную delta, равную 30 дням, для сравнения с ней 
    delta = timedelta(days= sheet_certdays_b2)

    # Создаем пустые списки, в которых будут аккумулироваться сообщения об ежегодном контроле
    message_1 = []
    message_2 = []
    message_3 = []

    i = 0 # Счетчик ячеек в столбце "C" файла Excel
    j = 0 # Счетчик ячеек на случай, если срок ежегодного контроля еще не подошел
    n = 0 # Счетчик пустых ячеек
    k = 0 # Счетчик ячеек на случай, если срок ежегодного контроля уже подошел и при этом его уже провели(стоит 1 в столбце "Е")
    
    # Запускаем цикл по ячейкам в столбце "C" Excel
    for cell_c in sheet["C"]:
        i +=1
        if (cell_c.value == "Дата выдачи аттестата"):
            continue
        if cell_c.value == None:
            n +=1
            continue
        # Условие 1, когда срок ежегодного контроля еще не подошел
        if ((datetime.now() - cell_c.value) < delta):
            j +=1
        # Условие 2, когда срок ежегодного контроля подошел
        if ((datetime.now() - cell_c.value) >= delta) and ((sheet.cell(row = i, column = 5)).value) != 1:
            message_3.append(f"\n У объекта {(sheet.cell(row = i, column = 2)).value} необходимо провести ежегодный контроль (с момента аттестации прошло более {delta.days} дней)")
        # Условие 3, когда срок ежегодного контроля уже подошел и при этом его уже провели(стоит 1 в столбце "Е")
        if ((datetime.now() - cell_c.value) >= delta) and ((sheet.cell(row = i, column = 5)).value) == 1:
            k +=1
         
    # Условие, когда нет объектов у которых необходимо провести ежегодный контроль
    if (j+n) == len(sheet["C"])-1:
        message_1.append(f"Нет объектов для проведения ежегодного контроля") 
        changed_message_1 = " ".join(message_1) 
        msgbox(changed_message_1,"ЦИТСиЗИ МВД по Республике Калмыкия")

    # Условие, когда срок ежегодного контроля уже подошел и при этом его уже провели(стоит 1 в столбце "Е")
    elif ((k+n+j) == len(sheet["C"])-1) and not ((j+n) == len(sheet["C"])-1):
        message_2.append("Ежегодный контроль проведен на всех объектах")  
        changed_message_2 = " ".join(message_2) 
        msgbox(changed_message_2,"ЦИТСиЗИ МВД по Республике Калмыкия") 
    # Условие, когда есть объекты для проведения ежегодного контроля 
    else:
        changed_message_3 = " ".join(message_3)     
        msgbox(changed_message_3,"ЦИТСиЗИ МВД по Республике Калмыкия")
